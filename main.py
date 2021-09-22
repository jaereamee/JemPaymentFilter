import os
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook #lib for managing excel https://openpyxl.readthedocs.io/en/stable/index.html 
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas.core.dtypes.missing import notnull

cwd = os.getcwd()

NRIC_COLUMN = "Please enter your NRIC/ Passport"
BANK_ACCOUNT_COLUMN = "Please enter Bank Account Holder's name" 

#######################################
# Bank Sheet Columns

PAYNOW_COLUMN = "Transaction Description 2"
AMOUNT_COLUMN = "Credit"

def startMenu():

    print("Please confirm the EXACT column names.")
    print("Full NRIC Column: "+NRIC_COLUMN)
    print("Bank Account Holder Name: "+BANK_ACCOUNT_COLUMN)
    confirm = input("Is this correct? (Y/N)")
    while True:
        if confirm.lower() == "y":
            smt
            break
        elif confirm.lower() == "n":
            no
            print("Which one needs to be changed?")
            for 
        else:
            print("Please confirm the EXACT column names.")
            print("Full NRIC Column: "+NRIC_COLUMN)
            print("Bank Account Holder Name: "+BANK_ACCOUNT_COLUMN)

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def bank_GetData():
    
    ## open the workbook, create if doesn't exist
    BANK_PATH = cwd+"/bankStatements.xlsx"
    if not os.path.isfile(BANK_PATH): 
        wb=Workbook()
        ws1 = wb.active
        ws1.title = "Raw"
        wb.save(filename = BANK_PATH)
        return 0
    
    wb_panda = pd.read_excel(BANK_PATH,engine="openpyxl",skiprows=5,dtype=str)
    wb_panda = wb_panda.fillna('')

    return wb_panda

def gf_GetData():
    """Get the Google Form Responses. It will be multiple column, with the key based on the fields of the form. Make this dynamic? Since google forms produces the first row as a fixed title name anyway"""
    GF_PATH = cwd+"/form2Responses.xlsx"

    # if the workbook doesnt exist create it. Ensure formatting is correct as well
    wb=Workbook()
    if not os.path.isfile(GF_PATH): 
        ws1 = wb.active
        ws1.title = "Raw"
        wb.create_sheet(title="Paid")
        wb.save(filename = GF_PATH)
        print("Please input the data")
        return 0
    # wb.active.title="Raw"
    # wb.save(filename=GF_PATH)
    
    # try again with pandas
    wb_panda = pd.read_excel(GF_PATH,sheet_name="Raw",engine="openpyxl", dtype=str)
    wb_panda = wb_panda.fillna('') # incase i nid to replace the nan with a blank

    # print("gf:\n")
    # print(wb_panda.head())

    return wb_panda
            
def masterFilter(GF,Bank):
    """
    The core filtering logic.
    Compare every entry in the gf with the bank statements. Store in a list of possible matches, if more than one remains, unlock the next layer of search. 
    Do in this order..
    1. Check Full IC
    2. Check first name segment from bank account
    3. If multiple, check next name segment among survivors
    4. Still multiple, check if GF has multiple same name
    """

    # set output file
    output_path = outputFileName()
    matches =[]
    bank_transaction=[]

    # lists to hold the filtered rows
    paid_too_many_times=[]
    paid_for_others=[]
    no_payment_found=[]
    paid_correctly=[]
    paid_wrong_amount=[]

    printProgressBar(0,len(GF.index),prefix = 'Progress:', suffix = 'Complete', length = 50)
    for a in len(GF):
        printProgressBar(a + 1, len(GF.index), prefix = 'Progress:', suffix = 'Complete', length = 50)
        gf_nric = GF[NRIC_COLUMN][a]
        name_list = GF[BANK_ACCOUNT_COLUMN][a]
        
        if search_NRIC(gf_nric,Bank):
            print(str(GF[BANK_ACCOUNT_COLUMN][a])+" found by nric")
            bank_duplicate, bt = checkAgainstBank(name_list,Bank)
            if bt>-1:
                bank_transaction.append(bt)
            if bank_duplicate:
                print(str(GF[BANK_ACCOUNT_COLUMN][a])+" - Paid too many times")
                paid_too_many_times.append(a)
            else:
                matches.append(a)
        elif search_Name(name_list,Bank):
            print(str(GF[BANK_ACCOUNT_COLUMN][a])+" found by name")
            bank_duplicate, bt = checkAgainstBank(name_list,Bank)
            if bt>-1:
                bank_transaction.append(bt)
            if bank_duplicate:
                if checkAgainstGF(GF,a):
                    print(str(GF[BANK_ACCOUNT_COLUMN][a])+" - Paid for others")
                    paid_for_others.append(a)
                else:
                    print(str(GF[BANK_ACCOUNT_COLUMN][a])+" - Paid too many times")
                    paid_too_many_times.append(a)
            else:
                matches.append(a)
        else: 
            print(str(GF[BANK_ACCOUNT_COLUMN][a])+" - No payment found")
            no_payment_found.append(a)
        a+=1
    print("matches: "+str(matches))
    ok=0
    
    for b in matches:
        # if (Bank.iloc[bank_transaction[b],PAYNOW_COLUMN]) == NULL:
        #     print(str(Bank.iloc[bank_transaction[b],0]) + " continue")
        print(str(Bank[AMOUNT_COLUMN][bank_transaction[b]]))
        
        if "98" in Bank[AMOUNT_COLUMN][bank_transaction[b]]:
            print(str(GF[BANK_ACCOUNT_COLUMN][b])+" OK")
            paid_correctly.append(b)
            ok = 1
            break
    if ok==0:
        print(str(GF[BANK_ACCOUNT_COLUMN][b])+" paid wrong amount")
        paid_wrong_amount.append(b)
            
            # else: print(str(GF[BANK_ACCOUNT_COLUMN][b])+"Paid wrong amount")
            # else: print(str(GF[BANK_ACCOUNT_COLUMN][b])+"Technical Error: can't find amount paid")
    
    return paid_too_many_times, paid_for_others, no_payment_found, paid_correctly, paid_wrong_amount



def search_NRIC(gf_nric, Bank):
    
    charnum=len(gf_nric)
    if charnum>0:
        for bah in len(Bank[PAYNOW_COLUMN]):
            if gf_nric.lower() in Bank[PAYNOW_COLUMN][bah].lower(): # if it matches any down the paynow column of the bank statements, return true
                return 1
        else:
            return 0


def search_Name(name_list, Bank):
    
    for bah in len(Bank[PAYNOW_COLUMN]): 
        if name_list.lower() in Bank[PAYNOW_COLUMN][bah].lower():
            return 1
    else:
        return 0

def checkAgainstBank(name_list,Bank):
    count=0
    address=999

    for bah in len(Bank[PAYNOW_COLUMN]):
        if name_list.lower() in Bank[PAYNOW_COLUMN][bah].lower():
            count+=1
            if count==1:
                address=bah
    if count>1:
        return 1,-1
    elif count==1:
        return 0,address
    else:
        return 0,0 # no record at all, which is impossible

def checkAgainstGF(GF,r):
    q=0
    count=0

    print(GF[BANK_ACCOUNT_COLUMN][r].lower())
    for q in len(GF):
        if GF[BANK_ACCOUNT_COLUMN][r].lower() in GF[BANK_ACCOUNT_COLUMN][q].lower():
            count+=1
        q+=1
    if count>1:
        return 1
    else:
        return 0
            
def ignoreCompleted(resultz,mylist):
    for d in len(resultz):
        if mylist[d] in resultz


def outputFileName():
    """Before the start of your printing, create a new output file that doesn't overwrite the existing (even those made in the same day)"""
    # check if the default name exists
        #if so, then iterate on the increments until you find one that's free. Set the output as that
    now = datetime.now().strftime("%Y%m%d")
    x=0
    output_file = now+"_output_"+str(x)
    OUTPUT_PATH = cwd+"/"+output_file+".xlsx"
    while 1:
        if os.path.isfile(OUTPUT_PATH):
            x+=1
            output_file = now+"_output_"+str(x)
            OUTPUT_PATH = cwd+"/"+output_file+".xlsx"
        else:
            break
    wb=Workbook()
    ws1 = wb.active
    ws1.title = "Paid Correctly"
    wb.create_sheet(title="Paid too many times")
    wb.create_sheet(title="Paid for others")
    wb.create_sheet(title="No payment found")
    wb.create_sheet(title="Paid wrong amount")
    wb.save(filename=OUTPUT_PATH)
    wb.close()

    return OUTPUT_PATH

def printMatches(GF,data,sheet,output_path):

    mylist =[]

    for i in data:
        mylist.append(GF.loc[i])

    # # if the workbook doesnt exist create it. Ensure formatting is correct as well
    wb=load_workbook(output_path)
    # ws = wb[sheet]
    writer=pd.ExcelWriter(output_path, engine='openpyxl')
    writer.book = wb

    #  check for duplicates..
    compares = pd.read_excel(output_path, 'Email Sent',dtype=str)
    compares = compares.fillna('')

    for j in len(df):
        if mylist[NRIC_COLUMN][j] not in compares[NRIC_COLUMN]:
            df2 

    
    df = pd.DataFrame(mylist,columns=GF.columns) # new dataframe tohold the list
        
    # finally write to the excel
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    df.to_excel(writer,sheet)
    writer.save()
    



if __name__ == "__main__":
    
    # x = gf_GetData()
    # y = bank_GetData()
    # paid_too_many_times, paid_for_others, no_payment_found, paid_correctly, paid_wrong_amount = masterFilter(x,y)
    # # search_Name(x,y,search_NRIC(x,y))
   
    # mypath = outputFileName()
    # print(mypath)

    # printMatches(x,paid_correctly,"Paid Correctly",mypath)
    # printMatches(x,paid_too_many_times,"Paid too many times",mypath)
    # printMatches(x,paid_for_others,"Paid for others",mypath)
    # printMatches(x,no_payment_found,"No payment found",mypath)
    # printMatches(x,paid_wrong_amount,"Paid wrong amount",mypath)

    ##########################################################################333

    he = bank_GetData()
    print(he.head())

